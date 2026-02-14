#!/usr/bin/env python3
import sys
import os
import time
import re

def convert_pdf_to_excel(input_path, output_path):
    try:
        print(f"Starting conversion: {input_path} -> {output_path}")
        start_time = time.time()

        import pymupdf as fitz
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        pdf_document = fitz.open(input_path)
        workbook = Workbook()

        total_pages = len(pdf_document)

        for page_num in range(total_pages):
            print(f"Processing page {page_num + 1}/{total_pages}")
            page = pdf_document.load_page(page_num)

            if page_num == 0:
                ws = workbook.active
                ws.title = f"Page 1"
            else:
                ws = workbook.create_sheet(title=f"Page {page_num + 1}")

            tables = page.find_tables()
            current_row = 1

            if tables and len(tables.tables) > 0:
                for t_idx, table in enumerate(tables.tables):
                    data = table.extract()

                    if t_idx > 0:
                        current_row += 1

                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    for row_idx, row_data in enumerate(data):
                        for col_idx, cell_value in enumerate(row_data):
                            cell = ws.cell(row=current_row + row_idx, column=col_idx + 1)
                            text = cell_value if cell_value else ""
                            text = text.strip() if isinstance(text, str) else str(text) if text is not None else ""

                            try:
                                num = float(text.replace(',', ''))
                                if num == int(num):
                                    cell.value = int(num)
                                else:
                                    cell.value = num
                            except (ValueError, AttributeError):
                                cell.value = text

                            cell.border = thin_border
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

                            if row_idx == 0:
                                cell.font = Font(bold=True, size=11)
                                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                            else:
                                cell.font = Font(size=11)

                            col_letter = get_column_letter(col_idx + 1)
                            current_width = ws.column_dimensions[col_letter].width or 8
                            text_len = len(str(cell.value)) + 2 if cell.value else 8
                            new_width = min(max(text_len, current_width), 50)
                            ws.column_dimensions[col_letter].width = new_width

                    current_row += len(data) + 1

                non_table_text = page.get_text("text")
                table_texts = set()
                for table in tables.tables:
                    for row in table.extract():
                        for cell in row:
                            if cell:
                                table_texts.add(cell.strip())

                remaining_lines = []
                for line in non_table_text.split('\n'):
                    line_stripped = line.strip()
                    if line_stripped and line_stripped not in table_texts:
                        remaining_lines.append(line_stripped)

                if remaining_lines:
                    current_row += 1
                    for line in remaining_lines:
                        cell = ws.cell(row=current_row, column=1)
                        cell.value = line
                        cell.font = Font(size=11)
                        current_row += 1

            else:
                text = page.get_text("text")
                if text.strip():
                    lines = text.split('\n')
                    for line in lines:
                        if line.strip():
                            parts = re.split(r'\s{2,}|\t', line.strip())

                            if len(parts) > 1:
                                for col_idx, part in enumerate(parts):
                                    cell = ws.cell(row=current_row, column=col_idx + 1)
                                    part = part.strip()
                                    try:
                                        num = float(part.replace(',', ''))
                                        if num == int(num):
                                            cell.value = int(num)
                                        else:
                                            cell.value = num
                                    except (ValueError, AttributeError):
                                        cell.value = part
                                    cell.font = Font(size=11)
                            else:
                                cell = ws.cell(row=current_row, column=1)
                                cell.value = line.strip()
                                cell.font = Font(size=11)

                            current_row += 1

                    for col in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        max_len = 0
                        for row in range(1, ws.max_row + 1):
                            val = ws.cell(row=row, column=col).value
                            if val:
                                max_len = max(max_len, len(str(val)))
                        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
                else:
                    cell = ws.cell(row=1, column=1)
                    cell.value = "No extractable text on this page (may be image-based)"
                    cell.font = Font(italic=True, color="999999")

        pdf_document.close()

        empty_sheets = [ws for ws in workbook.worksheets if ws.max_row <= 1 and not ws.cell(1, 1).value]
        if len(empty_sheets) < len(workbook.worksheets):
            for ws in empty_sheets:
                workbook.remove(ws)

        workbook.save(output_path)

        duration = round(time.time() - start_time, 2)
        print(f"Conversion completed in {duration} seconds")

        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            print(f"Output file created: {output_path} ({file_size} bytes)")
            return True
        else:
            print("Error: Output file not created")
            return False

    except Exception as e:
        print(f"Conversion error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python pdf_to_excel_converter.py <input_pdf> <output_xlsx>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    if not os.path.exists(input_file):
        print(f"Error: Input file not found: {input_file}")
        sys.exit(1)

    success = convert_pdf_to_excel(input_file, output_file)
    sys.exit(0 if success else 1)
